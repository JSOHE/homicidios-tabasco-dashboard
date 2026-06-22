# ============================================================
# PROYECTO: Homicidios en Tabasco con datos del SESNSP/RNID
# SCRIPT 05: Reporte automático en Excel
#
# OBJETIVO:
#   Crear un archivo Excel entregable con:
#   1. Portada
#   2. Resumen ejecutivo
#   3. Tablas analíticas
#   4. Ranking municipal
#   5. Promedio diario
#   6. Modalidades
#   7. Sexo y edad
#   8. Tentativas
#   9. Índice de gráficas
#   10. Gráficas insertadas como imágenes
#
# REQUISITO:
#   Antes de correr este script, debes haber corrido:
#   - Script 03: genera tablas analíticas
#   - Script 04: genera gráficas analíticas
#
# ENTRADAS:
#   outputs/tables/analiticas/
#   outputs/figures/analiticas/
#
# SALIDA:
#   outputs/reportes/reporte_homicidios_tabasco_2026_actual.xlsx
# ============================================================


# ------------------------------------------------------------
# 1. Importar librerías
# ------------------------------------------------------------

from pathlib import Path
from datetime import datetime

import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image


# ------------------------------------------------------------
# 2. Definir rutas del proyecto
# ------------------------------------------------------------

BASE_DIR = Path("/Users/soto/homicidios_tabasco")

TABLES_DIR = BASE_DIR / "outputs" / "tables" / "analiticas"
FIGURES_DIR = BASE_DIR / "outputs" / "figures" / "analiticas"
REPORTS_DIR = BASE_DIR / "outputs" / "reportes"

REPORTS_DIR.mkdir(parents=True, exist_ok=True)

SALIDA_EXCEL = REPORTS_DIR / "reporte_homicidios_tabasco_2026_actual.xlsx"


# ------------------------------------------------------------
# 3. Funciones auxiliares para leer tablas
# ------------------------------------------------------------

def leer_tabla(nombre_archivo):
    """
    Lee una tabla CSV desde outputs/tables/analiticas.

    Si no encuentra el archivo, muestra un error claro.
    Esto ayuda a saber si primero falta correr el script 03.
    """
    ruta = TABLES_DIR / nombre_archivo

    if not ruta.exists():
        raise FileNotFoundError(
            f"No encontré la tabla: {ruta}\n"
            "Primero corre el script 03."
        )

    return pd.read_csv(ruta)


def ruta_figura(nombre_archivo):
    """
    Devuelve la ruta de una imagen generada por el script 04.

    Si la imagen no existe, no detiene todo el programa:
    simplemente regresa None.
    """
    ruta = FIGURES_DIR / nombre_archivo

    if ruta.exists():
        return ruta

    print(f"Advertencia: no encontré la gráfica {ruta}")
    return None


# ------------------------------------------------------------
# 4. Funciones auxiliares de formato Excel
# ------------------------------------------------------------

def aplicar_formato_tabla(ws):
    """
    Aplica formato básico a una hoja de Excel:
    - encabezados en azul oscuro
    - texto blanco en encabezados
    - filtros
    - congelar primera fila
    - ajustar ancho de columnas
    - bordes suaves
    """

    azul = "1F4E78"
    blanco = "FFFFFF"
    gris_claro = "D9EAF7"
    borde_color = "B7B7B7"

    header_fill = PatternFill("solid", fgColor=azul)
    header_font = Font(color=blanco, bold=True)
    thin_border = Border(
        left=Side(style="thin", color=borde_color),
        right=Side(style="thin", color=borde_color),
        top=Side(style="thin", color=borde_color),
        bottom=Side(style="thin", color=borde_color),
    )

    # Si la hoja está vacía, no hacemos nada.
    if ws.max_row < 1 or ws.max_column < 1:
        return

    # Formato de encabezados.
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    # Formato del resto de celdas.
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # Congelar encabezado.
    ws.freeze_panes = "A2"

    # Activar autofiltro.
    ws.auto_filter.ref = ws.dimensions

    # Ajustar anchos de columnas de forma aproximada.
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_length = 0

        for cell in ws[col_letter]:
            valor = cell.value
            if valor is not None:
                max_length = max(max_length, len(str(valor)))

        adjusted_width = min(max(max_length + 2, 12), 35)
        ws.column_dimensions[col_letter].width = adjusted_width

    # Formato para números y porcentajes.
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            header = ws.cell(row=1, column=cell.column).value

            if header is None:
                continue

            header_text = str(header).lower()

            if "pct" in header_text or "porcentaje" in header_text:
                cell.number_format = "0.00"

            elif "promedio" in header_text:
                cell.number_format = "0.00"

            elif isinstance(cell.value, (int, float)):
                cell.number_format = "0"


def crear_portada(wb, resumen_control, estatal_mensual, ranking_municipal):
    """
    Crea la hoja de portada del reporte.

    Esta hoja resume:
    - fuente
    - periodo
    - principales indicadores
    - municipio con más víctimas
    """

    ws = wb.create_sheet("00_Portada", 0)

    # Colores
    azul = "1F4E78"
    azul_claro = "D9EAF7"
    blanco = "FFFFFF"
    gris = "666666"

    # Ajustar anchos principales.
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 22

    # Título principal
    ws.merge_cells("B2:E3")
    cell = ws["B2"]
    cell.value = "Reporte de Homicidios y Feminicidio\nTabasco, enero-mayo 2026"
    cell.font = Font(size=18, bold=True, color=blanco)
    cell.fill = PatternFill("solid", fgColor=azul)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Subtítulo
    ws.merge_cells("B5:E5")
    ws["B5"] = "Fuente: SESNSP / RNID - Víctimas municipal 2026"
    ws["B5"].font = Font(size=11, italic=True, color=gris)

    ws.merge_cells("B6:E6")
    ws["B6"] = f"Fecha de generación del reporte: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["B6"].font = Font(size=11, color=gris)

    # Extraer indicadores desde resumen_control
    def obtener_indicador(nombre):
        fila = resumen_control[resumen_control["indicador"] == nombre]
        if fila.empty:
            return 0
        return fila["victimas"].iloc[0]

    homicidio_doloso = obtener_indicador("homicidio_doloso")
    homicidio_culposo = obtener_indicador("homicidio_culposo")
    feminicidio = obtener_indicador("feminicidio")
    total_amplio = obtener_indicador("total_homicidios_y_feminicidio")
    promedio_periodo = obtener_indicador("promedio_diario_estatal_periodo")

    # Municipio con mayor total amplio
    ranking_ordenado = ranking_municipal.sort_values(
        "total_homicidios_y_feminicidio",
        ascending=False
    )

    municipio_top = ranking_ordenado.iloc[0]["Municipio"]
    victimas_top = ranking_ordenado.iloc[0]["total_homicidios_y_feminicidio"]
    pct_top = ranking_ordenado.iloc[0]["pct_total_amplio"]

    # Tabla de indicadores
    ws["B9"] = "Indicador"
    ws["C9"] = "Valor"

    indicadores = [
        ("Homicidio doloso", homicidio_doloso),
        ("Homicidio culposo", homicidio_culposo),
        ("Feminicidio", feminicidio),
        ("Total homicidios y feminicidio", total_amplio),
        ("Promedio diario estatal del periodo", promedio_periodo),
        ("Municipio con mayor total", municipio_top),
        ("Víctimas del municipio con mayor total", victimas_top),
        ("% estatal del municipio con mayor total", pct_top),
    ]

    fila_inicio = 10
    for i, (indicador, valor) in enumerate(indicadores, start=fila_inicio):
        ws[f"B{i}"] = indicador
        ws[f"C{i}"] = valor

    # Formato tabla indicadores
    for cell in ws["B9:C9"][0]:
        cell.fill = PatternFill("solid", fgColor=azul)
        cell.font = Font(color=blanco, bold=True)
        cell.alignment = Alignment(horizontal="center")

    for row in ws.iter_rows(min_row=10, max_row=17, min_col=2, max_col=3):
        for cell in row:
            cell.fill = PatternFill("solid", fgColor=azul_claro)
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    ws["C14"].number_format = "0.00"
    ws["C17"].number_format = "0.00"

    # Nota metodológica
    ws.merge_cells("B20:E25")
    ws["B20"] = (
        "Nota metodológica:\n"
        "El indicador 'Total homicidios y feminicidio' suma víctimas de homicidio doloso, "
        "homicidio culposo y feminicidio consumado. Las tentativas se presentan por separado "
        "y no se suman al total de homicidios consumados. El indicador 'total intencional' "
        "corresponde a homicidio doloso + feminicidio."
    )
    ws["B20"].alignment = Alignment(wrap_text=True, vertical="top")
    ws["B20"].fill = PatternFill("solid", fgColor="F2F2F2")

    # Quitar líneas de cuadrícula visuales en Excel.
    ws.sheet_view.showGridLines = False


def insertar_graficas(wb):
    """
    Crea una hoja llamada '12_Graficas' e inserta las imágenes
    generadas por el script 04.

    openpyxl inserta imágenes usando Image(ruta) y ws.add_image().
    """

    ws = wb.create_sheet("12_Graficas")

    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 45
    ws.column_dimensions["D"].width = 45
    ws.column_dimensions["E"].width = 45

    ws["B2"] = "Gráficas analíticas"
    ws["B2"].font = Font(size=16, bold=True, color="1F4E78")

    graficas = [
        (
            "01_estatal_barras_apiladas_homicidio_total.png",
            "Evolución estatal mensual del homicidio total consumado"
        ),
        (
            "02_promedio_diario_estatal_homicidio_total.png",
            "Promedio diario estatal por mes"
        ),
        (
            "03_ranking_municipal_total_amplio.png",
            "Ranking municipal del total amplio"
        ),
        (
            "04_ranking_municipal_total_intencional.png",
            "Ranking municipal de violencia letal intencional"
        ),
        (
            "05_heatmap_municipio_mes_total_amplio.png",
            "Heatmap municipio-mes del total amplio"
        ),
        (
            "06_modalidad_homicidio_doloso.png",
            "Modalidad del homicidio doloso"
        ),
        (
            "07_modalidad_homicidio_culposo.png",
            "Modalidad del homicidio culposo"
        ),
        (
            "08_sexo_edad_consumado.png",
            "Víctimas por sexo y edad"
        ),
        (
            "09_tentativas_estatales_mes.png",
            "Tentativas estatales por mes"
        ),
        (
            "10_promedio_diario_municipal_periodo.png",
            "Promedio diario municipal del periodo"
        ),
    ]

    fila = 4

    for archivo, titulo in graficas:
        ruta = ruta_figura(archivo)

        ws[f"B{fila}"] = titulo
        ws[f"B{fila}"].font = Font(size=12, bold=True, color="1F4E78")

        if ruta is not None:
            img = Image(str(ruta))

            # Ajustamos tamaño para que quepa en Excel.
            # Esto no modifica el archivo original, solo cómo se ve dentro del Excel.
            img.width = 760
            img.height = 430

            ws.add_image(img, f"B{fila + 1}")

            # Dejamos espacio para la siguiente gráfica.
            fila += 25
        else:
            ws[f"B{fila + 1}"] = "No se encontró la imagen. Corre primero el script 04."
            fila += 4

    ws.sheet_view.showGridLines = False


# ------------------------------------------------------------
# 5. Leer tablas analíticas
# ------------------------------------------------------------

resumen_control = leer_tabla("00_resumen_control.csv")
base_letal_larga = leer_tabla("01_base_letal_larga.csv")
base_consumada_larga = leer_tabla("01b_base_consumada_larga.csv")
municipal_mensual = leer_tabla("02_municipal_mensual_consumado.csv")
estatal_mensual = leer_tabla("03_estatal_mensual_consumado.csv")
ranking_municipal = leer_tabla("04_ranking_municipal_acumulado.csv")
heatmap_tabla = leer_tabla("05_heatmap_municipio_mes_total_amplio.csv")
modalidad_doloso = leer_tabla("06_modalidad_homicidio_doloso.csv")
modalidad_culposo = leer_tabla("07_modalidad_homicidio_culposo.csv")
sexo_edad = leer_tabla("08_sexo_edad_consumado.csv")
tentativas_municipal = leer_tabla("09_tentativas_municipal_mensual.csv")
tentativas_estatal = leer_tabla("10_tentativas_estatal_mensual.csv")
promedio_diario_municipal = leer_tabla("11_promedio_diario_municipal_mensual.csv")
promedio_diario_estatal = leer_tabla("12_promedio_diario_estatal_mensual.csv")


# ------------------------------------------------------------
# 6. Crear resumen ejecutivo como tabla
# ------------------------------------------------------------

# Esta tabla se genera para tener una hoja sintética adicional.

def obtener_indicador(nombre):
    fila = resumen_control[resumen_control["indicador"] == nombre]
    if fila.empty:
        return 0
    return fila["victimas"].iloc[0]


ranking_ordenado = ranking_municipal.sort_values(
    "total_homicidios_y_feminicidio",
    ascending=False
)

municipio_top = ranking_ordenado.iloc[0]["Municipio"]
victimas_top = ranking_ordenado.iloc[0]["total_homicidios_y_feminicidio"]
pct_top = ranking_ordenado.iloc[0]["pct_total_amplio"]

resumen_ejecutivo = pd.DataFrame({
    "Indicador": [
        "Homicidio doloso",
        "Homicidio culposo",
        "Feminicidio",
        "Total homicidios y feminicidio",
        "Total intencional",
        "Promedio diario estatal del periodo",
        "Municipio con mayor total amplio",
        "Víctimas del municipio con mayor total amplio",
        "Participación del municipio top en el total estatal (%)"
    ],
    "Valor": [
        obtener_indicador("homicidio_doloso"),
        obtener_indicador("homicidio_culposo"),
        obtener_indicador("feminicidio"),
        obtener_indicador("total_homicidios_y_feminicidio"),
        obtener_indicador("total_intencional"),
        obtener_indicador("promedio_diario_estatal_periodo"),
        municipio_top,
        victimas_top,
        pct_top
    ]
})


# ------------------------------------------------------------
# 7. Escribir tablas en Excel con pandas
# ------------------------------------------------------------

# pandas.ExcelWriter permite escribir varias hojas dentro del mismo archivo.
# Cada DataFrame se exporta a una hoja distinta.

with pd.ExcelWriter(SALIDA_EXCEL, engine="openpyxl") as writer:

    resumen_ejecutivo.to_excel(
        writer,
        sheet_name="01_Resumen",
        index=False
    )

    resumen_control.to_excel(
        writer,
        sheet_name="02_Control",
        index=False
    )

    estatal_mensual.to_excel(
        writer,
        sheet_name="03_Estatal_Mensual",
        index=False
    )

    promedio_diario_estatal.to_excel(
        writer,
        sheet_name="04_Prom_Diario_Est",
        index=False
    )

    ranking_municipal.to_excel(
        writer,
        sheet_name="05_Ranking_Mpal",
        index=False
    )

    promedio_diario_municipal.to_excel(
        writer,
        sheet_name="06_Prom_Diario_Mpal",
        index=False
    )

    municipal_mensual.to_excel(
        writer,
        sheet_name="07_Mpal_Mensual",
        index=False
    )

    heatmap_tabla.to_excel(
        writer,
        sheet_name="08_Heatmap_Tabla",
        index=False
    )

    modalidad_doloso.to_excel(
        writer,
        sheet_name="09_Mod_Doloso",
        index=False
    )

    modalidad_culposo.to_excel(
        writer,
        sheet_name="10_Mod_Culposo",
        index=False
    )

    sexo_edad.to_excel(
        writer,
        sheet_name="11_Sexo_Edad",
        index=False
    )

    tentativas_estatal.to_excel(
        writer,
        sheet_name="13_Tentativas_Est",
        index=False
    )

    tentativas_municipal.to_excel(
        writer,
        sheet_name="14_Tentativas_Mpal",
        index=False
    )

    # Las bases largas pueden ser grandes, por eso las dejamos al final.
    base_consumada_larga.to_excel(
        writer,
        sheet_name="15_Base_Consumada",
        index=False
    )

    base_letal_larga.to_excel(
        writer,
        sheet_name="16_Base_Letal",
        index=False
    )


# ------------------------------------------------------------
# 8. Abrir el archivo con openpyxl para darle formato
# ------------------------------------------------------------

wb = load_workbook(SALIDA_EXCEL)


# ------------------------------------------------------------
# 9. Crear portada
# ------------------------------------------------------------

crear_portada(
    wb=wb,
    resumen_control=resumen_control,
    estatal_mensual=estatal_mensual,
    ranking_municipal=ranking_municipal
)


# ------------------------------------------------------------
# 10. Aplicar formato a todas las hojas de tablas
# ------------------------------------------------------------

for ws in wb.worksheets:
    # No aplicamos formato de tabla a portada ni a hoja de gráficas.
    if ws.title not in ["00_Portada", "12_Graficas"]:
        aplicar_formato_tabla(ws)


# ------------------------------------------------------------
# 11. Insertar gráficas en una hoja nueva
# ------------------------------------------------------------

insertar_graficas(wb)


# ------------------------------------------------------------
# 12. Reordenar hojas
# ------------------------------------------------------------

# openpyxl maneja las hojas en wb._sheets.
# Vamos a dejar un orden lógico para el reporte.

orden_deseado = [
    "00_Portada",
    "01_Resumen",
    "02_Control",
    "03_Estatal_Mensual",
    "04_Prom_Diario_Est",
    "05_Ranking_Mpal",
    "06_Prom_Diario_Mpal",
    "07_Mpal_Mensual",
    "08_Heatmap_Tabla",
    "09_Mod_Doloso",
    "10_Mod_Culposo",
    "11_Sexo_Edad",
    "12_Graficas",
    "13_Tentativas_Est",
    "14_Tentativas_Mpal",
    "15_Base_Consumada",
    "16_Base_Letal",
]

hojas_ordenadas = []

for nombre in orden_deseado:
    if nombre in wb.sheetnames:
        hojas_ordenadas.append(wb[nombre])

# Si por alguna razón hay hojas extra, las agregamos al final.
for ws in wb.worksheets:
    if ws.title not in orden_deseado:
        hojas_ordenadas.append(ws)

wb._sheets = hojas_ordenadas


# ------------------------------------------------------------
# 13. Guardar archivo final
# ------------------------------------------------------------

wb.save(SALIDA_EXCEL)

print("\nReporte Excel generado correctamente:")
print(SALIDA_EXCEL)

print("\nProceso terminado correctamente.")